import json
from pathlib import Path

import openpyxl


def build_mapping(xlsx_path: str) -> dict:
    """
    Build a mapping of Police Station name -> list of areas/villages.

    - Trims extra spaces around names
    - Removes duplicate areas per station (case-insensitive)
    - Keeps the original spelling/case as in the Excel file
    """
    wb = openpyxl.load_workbook(xlsx_path)
    sh = wb[wb.sheetnames[0]]

    # Start from row 2 (row 1 is the title row).
    data_rows = list(sh.iter_rows(min_row=2, values_only=True))

    mapping: dict[str, list[str]] = {}
    current_ps: str | None = None

    # Skip the first data row, which contains the column headers like "S.No"
    for row in data_rows[1:]:
        s_no, ps_s_no, sub_div, ps_name, area = row

        # When a new PS name appears, update current_ps
        if isinstance(ps_name, str) and ps_name.strip():
            ps_name_clean = ps_name.strip()
            current_ps = ps_name_clean
            mapping.setdefault(current_ps, [])

        # If we still don't have a PS name, skip
        if not current_ps:
            continue

        # Read area / village name
        if not isinstance(area, str):
            continue
        area_clean = area.strip()
        if not area_clean:
            continue

        # Deduplicate per PS using a normalized key, but keep original spelling
        existing_norm = {a.strip().lower() for a in mapping[current_ps]}
        norm = area_clean.lower()
        if norm not in existing_norm:
            mapping[current_ps].append(area_clean)

    return mapping


def write_json(mapping: dict, path: Path) -> None:
    """Write mapping to a pretty JSON file (UTF-8)."""
    path.write_text(
        json.dumps(mapping, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_mapping_js(mapping: dict, path: Path) -> None:
    """
    Write mapping to a JavaScript file as:

        const policeStationAreas = {...};

    This file can be included via <script src="mapping.js"></script>.
    """
    with path.open("w", encoding="utf-8") as f:
        f.write("const policeStationAreas = ")
        json.dump(mapping, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")


if __name__ == "__main__":
    excel_path = "ATP Dist_All PSs Villages-Areas.xlsx"
    mapping = build_mapping(excel_path)

    # Write mapping.json (pretty, useful for inspection if needed)
    write_json(mapping, Path("mapping.json"))

    # Write mapping.js (compact, loaded directly by index.html)
    write_mapping_js(mapping, Path("mapping.js"))

    # Also print JSON to stdout for quick checks
    print(json.dumps(mapping, ensure_ascii=False, indent=2))

